#-*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
def group_smpl_by_content(true_smpl_name):
    smpl_contents=["olej", "wit", "AA", "sól", "mąka", "białko", "NPs", "mix k", "mix peps", "mix pank", "maka", "bialko", "sol", "aa"]
    for content in smpl_contents:
        if(content in true_smpl_name):
            return content


        wb = Workbook()
        ws = wb.active
        for k in range (len(smpl_contents)):
            d=ws.cell(row=3, column=k+2, value=smpl_contents[k])
            d = ws.cell(row=3, column=k + 3 +len(smpl_contents), value=smpl_contents[k])
        wb.save("results.xlsx")

