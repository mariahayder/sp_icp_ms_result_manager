#-*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout
import os

from find_smpl_number import *
from find_true_smpl_name import *
from smpl_grouping_by_content import *
import string
#FILENAME='006SMPL_ParticleSizeDistribution-65Cu.xlsx'
#TABLE_FILENAME='table 08.06.xlsx'

def create_chart(FILENAME, TABLE_FILENAME):
    SMPL = find_smpl_number(FILENAME)


    true_smpl_name, median_size = find_true_smpl_name(TABLE_FILENAME, SMPL)

    #os.chdir("files")
    wb = load_workbook('%s' % FILENAME)
    ws = wb.active

    max_size=[]
    end = False
    current_row = 5
    final_row = 4
    rows = []
    beginning = True
    subsequent_zeros = 0
    while (end == False):
        if subsequent_zeros > 4:
            end = True
        d = ws.cell(row=current_row, column=2)
        if (d.value != None):
            if (d.value == 0 and beginning == True):
                current_row += 1
                continue
            else:
                beginning = False
                new_row = []
                e = ws.cell(row=current_row, column=1)
                new_row.append(e.value)
                new_row.append(d.value)
                rows.append(new_row)
                if(d.value==100):
                    max_size.append(e.value)
                final_row += 1
                current_row += 1
                if (d.value == 0):
                    subsequent_zeros += 1
        else:
            end = True
    print(final_row)
    print(rows)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2['A1'] = "size"
    ws2['B1'] = "frequency"
    for row in rows:
        ws2.append(row)

    chart1 = BarChart()
    # chart1.type = "col"
    chart1.style = 3
    chart1.title = '%s' % true_smpl_name
    chart1.y_axis.title = 'normalized frequency / %'
    chart1.x_axis.title = 'size / nm'

    data = Reference(ws2, min_col=2, min_row=2, max_row=final_row - 4, max_col=2)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=final_row - 4, max_col=1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 3
    chart1.legend = None
    chart1.y_axis.scaling.max = 100
    chart1.height = 15
    chart1.width = 25
    chart1.x_axis.delete = False
    chart1.y_axis.delete = False

    ws2.add_chart(chart1, "D10")
    print(cats)
    ws2['D3'] = "median size / nm"
    ws2['D4']=median_size
    FILENAME=FILENAME.rstrip('.xlsx')
    wb2.save("{}_{}_CHART.xlsx".format(FILENAME, true_smpl_name))
    true_smpl_name=group_smpl_by_content(true_smpl_name)
    size=[max_size, median_size]
    return true_smpl_name, size





