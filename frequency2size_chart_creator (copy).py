from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout

wb=load_workbook('SMPL.xlsx')
ws=wb.active

end=False
current_row=5
final_row=4
rows=[]
while(end==False):
    d=ws.cell(row=current_row, column=1)
    if(d.value!=None):
        new_row=[]
        e=ws.cell(row=current_row, column=2)
        new_row.append(d.value)
        new_row.append(e.value)
        rows.append(new_row)
        final_row+=1
        current_row+=1
    else:
        end=True
print(final_row)
print(rows)

wb2=Workbook()
ws2=wb2.active
ws2['A1']="size"
ws2['B1']="frequency"
for row in rows:
    ws2.append(row)


chart1 = BarChart()
#chart1.type = "col"
chart1.style = 3
chart1.title = "Bar Chart"
chart1.y_axis.title = 'normalized frequency / %'
chart1.x_axis.title = 'size / nm'

data = Reference(ws2, min_col=2, min_row=2, max_row=final_row-4, max_col=2)
cats = Reference(ws2, min_col=1, min_row=2, max_row=final_row-4, max_col=1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 3
chart1.legend = None
chart1.y_axis.scaling.max = 100
chart1.height=15
chart1.width=25
chart1.x_axis.delete = False
chart1.y_axis.delete = False

ws2.add_chart(chart1, "D10")
print(cats)


wb2.save("SMPL_COPY.xlsx")
