#-*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

SAMPLE_CONTENTS=10
def create_table(size_data, median_data):
    wb=load_workbook("results.xlsx")
    ws=wb.active
    d = ws.cell(row=2, column=2, value="most frequent size / nm")
    d = ws.cell(row=2, column=4+len(size_data), value="median size / nm")

    last_row_of_data=3
    #col=0
    for key in size_data:
        n=2
        for n in range(2, 2+SAMPLE_CONTENTS):
            if((ws.cell(row=3, column=n).value)==key):
                break
        max_sizes=size_data[key]
        for i in range(len(max_sizes)):
            d = ws.cell(row=i+4, column=n, value=max_sizes[i])
        if(last_row_of_data<len(max_sizes)+3):
            last_row_of_data=len(max_sizes)+3
            print("last row of data{}".format(last_row_of_data))

    for key in median_data:
        n=SAMPLE_CONTENTS+2
        for n in range(2+SAMPLE_CONTENTS, 2*SAMPLE_CONTENTS+2):
            if((ws.cell(row=3, column=n).value)==key):
                break
        medians=median_data[key]
        letter = get_column_letter(n)
        d = ws.cell(row=last_row_of_data + 1, column=n, value='=AVERAGE({}{}:{}{})'.format(letter, 4, letter, last_row_of_data))
        d = ws.cell(row=last_row_of_data + 2, column=n, value='=STDEV({}{}:{}{})'.format(letter, 4, letter, last_row_of_data))
        for i in range(len(medians)):
            d = ws.cell(row=i+4, column=n, value=medians[i])


    for key in size_data:
        n=2
        for n in range(2, 2+SAMPLE_CONTENTS):
            if((ws.cell(row=3, column=n).value)==key):
                break
        letter = get_column_letter(n)
        d=ws.cell(row=last_row_of_data+1, column=n, value='=AVERAGE({}{}:{}{})'.format(letter, 4, letter, last_row_of_data))
        d = ws.cell(row=last_row_of_data + 1, column=1, value="average")
        d = ws.cell(row=last_row_of_data + 2, column=n, value='=STDEV({}{}:{}{})'.format(letter, 4, letter, last_row_of_data))
        d = ws.cell(row=last_row_of_data + 2, column=1, value="standard deviation")
    wb.save("results.xlsx")


