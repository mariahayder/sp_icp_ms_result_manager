from openpyxl import Workbook, load_workbook

def find_true_smpl_name(TABLE_FILENAME, SMPL):
    wbt = load_workbook('%s' % TABLE_FILENAME)
    wst = wbt.active
    true_smpl_name = "name"
    should_break = False
    median_size=0
    for row in wst.iter_rows(min_row=3, min_col=3, max_col=3):
        for cell in row:
            print(cell.value)
            if (SMPL in cell.value):
                print('{}'.format(cell.row))
                true_smpl_name = wst.cell(cell.row, column=6).value
                median_size = wst.cell(cell.row, column=13).value
                should_break = True
                break
        if (should_break == True):
            break

    print(true_smpl_name)
    return true_smpl_name, median_size

def find_true_emf_name(TABLE_FILENAME, SMPL):
    print("table in find: {}".format(TABLE_FILENAME))
    wbt = load_workbook('%s' % TABLE_FILENAME)
    wst = wbt.active
    true_smpl_name = "name"
    should_break = False
    #median_size=0
    for row in wst.iter_rows(min_row=3, min_col=3, max_col=3):
        for cell in row:
            print(cell.value)
            if (SMPL in cell.value):
                print('{}'.format(cell.row))
                true_smpl_name = wst.cell(cell.row, column=6).value
                #median_size = wst.cell(cell.row, column=13).value
                should_break = True
                break
        if (should_break == True):
            break

    print(true_smpl_name)
    return true_smpl_name